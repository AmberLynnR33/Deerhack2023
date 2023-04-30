from openpyxl import Workbook, load_workbook, worksheet
from typing import Optional
from tkinter.filedialog import askopenfilename

class Model:
    """
    a class to represent an excel file
    === public attributes ===
    wb: the current workbook
    ws: the current worksheet (for the accessed month and date)
    cat: a dictionary of catagories and the amount spent on each
    amount_made: the amount of money the user made
    goals: this user's goals for the current month
    """
    wb: Workbook
    ws: worksheet
    cat: dict
    amount_made: float
    goals: dict

    def __init__(self) -> None:
        """
        initialize a new workbook
        """
        self.wb = Workbook()
        self.ws = None  # empty workbook initially
        self.cat = {'Essentials': 0, 'Bills': 0, 'Subscriptions': 0,
                    'Education / work': 0, 'Luxuries': 0}
        # initially all 0, since the user has not inputted anything yet
        self.amount_made = 0  # initially the user makes no money
        self.goals = {'Essentials': 0, 'Bills': 0, 'Subscriptions': 0,
                      'Education / work': 0, 'Luxuries': 0}
        # initialize view

    def new_file(self) -> None:
        """
        initializes a new file
        """
        self.wb = Workbook()
        self.ws = self.wb.active

    # otherwise, create new workbook
    def open_file(self) -> None:
        """
        opens an existing file
        return None if the path is invalid
        """
        self.file_name = self.save_file_path()
        if self.file_name is not None:  # handle error
            self.wb = load_workbook(f'{self.file_name}')  # each application is a sheet
        self.ws = self.wb.active

    def save_file(self) -> None:
        """
        saves the file
        """
        #file_name = self.save_file_path()
        self.wb.save(f'{self.file_name}')  # save workbook

    def save_file_path(self) -> Optional[str]:
        try:
            return askopenfilename()
        except FileNotFoundError:
            return None

    # we need to access the worksheets according to the month,
    # or create a new one
    # this is where we will input data
    def page_exists(self, month: int, year: int) -> None:
        '''check if the worksheet with this month and year exists'''
        if self.ws.title == 'Sheet':
            self.ws.title = f'{year},{month}'
        else:
            if f'{year},{month}' not in self.wb.sheetnames:
                self._new_page(month, year)
                self.ws = sheet
            else:
                # set current sheet to this found sheet
                for sheet in self.wb:
                    if sheet.title == f'{year},{month}':
                        self.ws = sheet

    def _new_page(self, month: int, year: int) -> None:
        '''
        initializes a new page in the excel
        '''
        # set current sheet to this new sheet
        self.ws = self.wb.create_sheet(f'{year},{month}')
        # add the column into the sheet, each catagory is a row
        for cat in self.cat:
            self.ws.append([cat])
        self.amount_made = 0
        for cat in self.cat:  # reset the amount spent
            self.cat[cat] = 0
        for goal in self.goals:  # reset goals
            self.goals[goal] = 0

    # *** USER SETTINGS *** #
    def add_money_out(self, amount: float, catagory: str) -> None:
        """
        add the amount spent to this specific catagory
        """
        catagory = catagory.capitalize()
        col_counter = 1  # counter for number of columns
        row_counter = 1  # counter for number of columns
        names = []
        # accumulate for total cost
        self.cat[catagory] += float(amount)
        if self.cat[catagory] > self.goals[catagory]:
            self.spent_too_much()
        # find the empty column for that row, and add the amount to it, test
        for row in self.ws.iter_rows():
            names.append(self.ws.cell(row_counter, 1).value)
            if self.ws.cell(row_counter, 1).value == catagory:  # found row
                for col in self.ws.iter_cols(min_col=2):
                    if (self.ws.cell(row_counter, col_counter+1)).value is None:
                        break
                    col_counter += 1  # count until we reach none
                self.ws.cell(row_counter, col_counter + 1, amount)
                break
            row_counter += 1
        if catagory not in names:
            self.ws.append([f'{catagory}', amount])

    def add_money_in(self, amount: float) -> None:
        """
        allows the user to add the amount of money they make
        """
        self.amount_made += float(amount)
        # add it to the spreadsheet, as a new row if amount made doesnt exist
        # get the names of the rows
        names = []
        row_counter = 1
        col_counter = 1
        for row in self.ws.iter_rows():
            names.append(self.ws.cell(row_counter, 1).value)
            if self.ws.cell(row_counter, 1).value == 'Amount Made':
                for col in self.ws.iter_cols(min_col=2):
                    if (self.ws.cell(row_counter, col_counter+1)).value is None:
                        break
                    col_counter += 1
                self.ws.cell(row_counter, col_counter, amount)
                break
            row_counter += 1
        if 'Amount Made' not in names:
            self.ws.append(['Amount Made', amount])

    def add_goal(self, amount: float, catagory: str) -> None:
        """
        allows the user to add in a goal for a specific catagory
        """
        catagory = catagory.capitalize()
        self.goals[catagory] = amount
        # want to add this goal as a row in the excel
        names = []
        row_counter = 1
        for row in self.ws.iter_rows():
            names.append(self.ws.cell(row_counter, 1).value)
            if self.ws.cell(row_counter, 1).value == f'{catagory} goal':  # goal
                # exists, just edit
                self._edit_goals(amount, catagory, row_counter)
                break
            row_counter += 1
        if f'{catagory} goal' not in names:  # add as a row
            self.ws.append([f'{catagory} goal', amount])

    def _edit_goals(self, amount: float, catagory: str, row: int) -> None:
        """
        allows the user to edit their goal for a specific catagory
        """
        self.ws.cell(row, 2, amount)

    def total_amount_earned(self) -> float:
        """
        returns the total amount of money this user earned
        """
        return self.amount_made

    def total_spent(self, catagory=None) -> float:
        """
        returns the total amount of money spent for the specific catagory
        """
        if catagory is not None:
            catagory = catagory.capitalize()
            return self.cat[catagory]
        else:
            c = 0  # counter, add up all the money
            for cat in self.cat:
                c += self.cat[cat]
            return c

    def spent_too_much(self) -> str:
        """
        return a message if the user spends too much
        """
        return "you spent too much money!"  # edit lol

    def get_goals(self) -> dict:
        """
        return a dictionary of this user's goals for each catagory
        """
        return self.goals
